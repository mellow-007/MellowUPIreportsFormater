import tkinter as tk
from tkinter import filedialog
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import os
from datetime import datetime
import zipfile
import win32com.client as win32
import win32print

# Prompt user to browse and select file(s)
root = tk.Tk()
root.withdraw()
filetypes = [("ZIP and CSV files", "*.zip;*.csv")]
file_paths = filedialog.askopenfilenames(title="Select file(s)", filetypes=filetypes)

processed_files = []  # Store the paths of processed files



# Process the selected files
for file_path in file_paths:
    # Get the base name without extension
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    
    # Check if the file is a zip file
    if file_path.lower().endswith('.zip'):
        # Extract the ZIP file in the same location
        zip_dir = os.path.dirname(file_path)
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            csv_files = [f for f in zip_ref.namelist() if f.lower().endswith('.csv')]
            if len(csv_files) == 0:
                continue  # Skip if no CSV file found
            extracted_files = []
            for csv_file in csv_files:
                extracted_path = os.path.join(zip_dir, csv_file)
                zip_ref.extract(csv_file, zip_dir)
                extracted_files.append(extracted_path)
            
            # Create a new folder with the same name as the ZIP file
            folder_name = base_name
            folder_number = 1
            while os.path.exists(os.path.join(zip_dir, folder_name)):
                folder_name = f"{base_name}_{folder_number}"
                folder_number += 1
            new_folder_path = os.path.join(zip_dir, folder_name)
            os.makedirs(new_folder_path)
            
            # Move the extracted CSV files to the new folder
            for extracted_file in extracted_files:
                new_file_path = os.path.join(new_folder_path, os.path.basename(extracted_file))
                os.rename(extracted_file, new_file_path)
            
            # Continue processing within the new folder
            os.chdir(new_folder_path)
            file_path = new_file_path
    
    # For non-zip files, process as a regular CSV file
    if file_path.lower().endswith('.csv'):
        # Create a new folder with the same name as the CSV file
        folder_name = base_name
        folder_number = 1
        while os.path.exists(os.path.join(os.path.dirname(file_path), folder_name)):
            folder_name = f"{base_name}_{folder_number}"
            folder_number += 1
        new_folder_path = os.path.join(os.path.dirname(file_path), folder_name)
        os.makedirs(new_folder_path)
        
        # Move the CSV file to the new folder
        new_file_path = os.path.join(new_folder_path, os.path.basename(file_path))
        os.rename(file_path, new_file_path)
        
        # Continue processing within the new folder
        os.chdir(new_folder_path)
        file_path = new_file_path

    # Append the processed file information to the list
    processed_files.append((file_path, file_path.lower().endswith('.csv')))

# Process the files
for file_path, is_csv in processed_files:
    if is_csv:
        # Use os.path.join to create the correct file path
        full_file_path = os.path.join(new_folder_path, os.path.basename(file_path))
        
        # Read the CSV file
        with open(full_file_path, 'r', newline='') as file:
            reader = csv.reader(file)
            rows = list(reader)
            
        # Reversing the rows from the third row onwards
        rows[2:] = reversed(rows[2:])



    # Clear column A and rename A1 as "S.No"
    rows[0][0] = 'S.No'
    for row in rows[1:]:
        if row:
            row[0] = ''

    # Delete the second row
    if len(rows) > 1:
        del rows[1]

    # Number the rows in column A starting from 1
    for i, row in enumerate(rows[1:], start=1):
        if row:
            row[0] = i

    # Remove columns B, C, D, F, K, L, M, N, O, P, Q, R, S
    columns_to_remove = [1, 2, 3, 5, 10, 11, 12, 13, 14, 15, 16, 17, 18]
    for row in rows:
        for col_index in sorted(columns_to_remove, reverse=True):
            if len(row) > col_index:
                del row[col_index]

    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active

    # Write the modified data to the worksheet
    for row in rows:
        worksheet.append(row)

    # Set the alignment and height of cells, and wrap text
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet.row_dimensions[cell.row].height = 15

    # Clear all cells in column F except F1
    for row in worksheet.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            if cell.row != 1:
                cell.value = ''

    # Set the width of columns
    column_widths = {'A': 5, 'B': 25, 'C': 9, 'D': 25, 'E': 11}
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    # Set the value of cell F1 to "Remarks"
    worksheet['F1'].value = "Remarks"

    # Set the width of column F to 9
    worksheet.column_dimensions['F'].width = 9

    # Set the font style to bold for cells A1, B1, C1, D1, E1, and F1
    bold_font = Font(bold=True)
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = worksheet[col + '1']
        cell.font = bold_font

    # Convert the date and time format in column D
    for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.value = cell.value.strftime("%d/%m/%Y %I:%M:%S %p")
            else:
                try:
                    datetime_obj = datetime.strptime(cell.value, "%Y-%m-%d %H:%M:%S")
                    formatted_date = datetime_obj.strftime("%d/%m/%Y %I:%M:%S %p")
                    cell.value = formatted_date
                except ValueError:
                    pass

    # Convert column C to whole number if decimal value is 0
    for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            try:
                value = float(cell.value)
                if value.is_integer():
                    cell.value = int(value)
            except ValueError:
                pass

    # Add a row at the end of each day based on day value in column D
    previous_day = None
    for row in reversed(list(worksheet.iter_rows(min_row=2, min_col=4, max_col=4))):
        for cell in row:
            if isinstance(cell.value, str) and len(cell.value) >= 2:
                current_day = int(cell.value[:2])
                if previous_day is not None and current_day != previous_day:
                    if cell.row + 1 <= worksheet.max_row:
                        worksheet.insert_rows(cell.row + 1)
                        previous_day = current_day
                elif previous_day is not None and current_day == previous_day:
                    if cell.row + 1 <= worksheet.max_row:
                        cell = worksheet.cell(row=cell.row + 1, column=cell.column)
                        previous_day = current_day
                else:
                    previous_day = current_day

    # Define the fill color
    orange_fill = PatternFill(start_color="FFFF9900", end_color="FFFF9900", fill_type="solid")
    gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    # Color the empty rows up to column F
    for row in worksheet.iter_rows():
        is_empty_row = all(cell.value is None for cell in row)
        if is_empty_row:
            for cell in row[:6]:  # Only color cells in columns A to F
                cell.fill = orange_fill

    # Find the last row of data
    last_row = worksheet.max_row

    # Color the rows after the last row of data up to column F
    for row in worksheet.iter_rows(min_row=last_row + 1, max_row=worksheet.max_row):
        is_empty_row = all(cell.value is None for cell in row[:6])  # Only check cells in columns A to F
        if is_empty_row:
            for cell in row[:6]:  # Only color cells in columns A to F
                cell.fill = orange_fill
        else:
            break

    # Set the width of column E to 11
    worksheet.column_dimensions['E'].width = 11

    # Set the height of all cells to 15
    for row in worksheet.iter_rows():
        for cell in row:
            worksheet.row_dimensions[cell.row].height = 15

    # Find the first row after the last row of data
    first_row_after_data = last_row + 1

    # Set the fill color for the first row after the data
    for cell in worksheet[first_row_after_data]:
        cell.fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
    
    # Color the even-numbered rows up to column F, excluding rows with existing color
    for row in worksheet.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=6):
        if row[0].row % 2 == 0:
            has_color = any(cell.fill != PatternFill(fill_type=None) for cell in row)
            if not has_color:
                for cell in row:
                    cell.fill = gray_fill

    # Set the value of cell C1 to "Amount"
    worksheet['C1'].value = "Amount"

    # Set the value of cell E1 to "Status"
    worksheet['E1'].value = "Status"

        # Save the modified file as an Excel workbook
    excel_file_path = os.path.splitext(file_path)[0] + '.xlsx'
    workbook.save(excel_file_path)

        # Open the modified file
    os.startfile(excel_file_path)

    # Convert Excel file to PDF
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    workbook = excel.Workbooks.Open(excel_file_path)
    pdf_file_path = os.path.splitext(file_path)[0] + '.pdf'
    workbook.ExportAsFixedFormat(0, pdf_file_path)
    workbook.Close()

    # Open the converted PDF file
    os.startfile(pdf_file_path)
